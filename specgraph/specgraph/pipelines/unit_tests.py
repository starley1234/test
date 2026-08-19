"""Пайплайн инженера верификации: тесты по требованию → Excel → прогон (если есть код).

Код прошивки в задаче не приложен: без него лист Simulation честно пишет
«прогон не выполнен». Если в запросе есть source_code — статическая проверка
(есть ли функция, опасные регистры), без исполнения на МК.
"""

from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path
from typing import Any
from uuid import uuid4

from sqlalchemy.orm import Session

from specgraph.config import settings
from specgraph.llm import chat_llm
from specgraph.models import Requirement
from specgraph.retrieval.context import expand_requirement

EXPORTS = Path(settings.upload_dir).resolve().parent / "exports"
PROMPT = Path(__file__).with_name("unit_tests_prompt.txt")


def _safe_name(code: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", code).strip("_")


def _guess_func(req: Requirement) -> str:
    blob = " ".join([req.code or "", req.text or "", req.title or ""])
    m = re.search(r"\b([a-zA-Z_][a-zA-Z0-9_]{2,})\s*\(\)", blob)
    if m:
        return m.group(1)
    m = re.search(r"\b(adcSetup|freqPhDef|dmaSetup|memcpy)\b", blob)
    return m.group(1) if m else "uut"


def _is_hw(req: Requirement) -> bool:
    blob = f"{req.code} {req.text} {req.kind.value}".lower()
    return any(x in blob for x in ("hwrq", "adc", "регистр", "dma", "gpio", "таймер", "peripheral"))


def _heuristic_suite(req: Requirement) -> dict[str, Any]:
    func = _guess_func(req)
    nums = [float(x.replace(",", ".")) for x in re.findall(r"-?\d+(?:[.,]\d+)?", req.text or "")]
    mid = nums[0] if nums else 0.0
    lo, hi = (min(nums), max(nums)) if nums else (-1.0, 1.0)
    hw = _is_hw(req)
    danger = (
        "ВЫСОКАЯ: запись в регистры/периферию — риск выхода МК из строя"
        if hw
        else "Низкая"
    )
    cases = [
        {
            "purpose": "Нормальные условия. Типовое значение из класса эквивалентности допустимых.",
            "method": "Классы эквивалентности",
            "actions": f"Вызвать {func}() с типовым входом {mid}. Проверить ожидаемое состояние.",
            "expect": "Функция отрабатывает без ошибки, выход в допустимом диапазоне.",
            "inputs": {"in0": mid, "state": 0},
            "outputs": {"out0": mid, "state": 1},
            "types": {"in0": "single", "state": "uint16", "out0": "single"},
        },
        {
            "purpose": "Нижняя граница диапазона.",
            "method": "Граничные значения",
            "actions": f"Вызвать {func}() при in0={lo}.",
            "expect": "Корректная обработка нижней границы.",
            "inputs": {"in0": lo, "state": 0},
            "outputs": {"out0": lo, "state": 1},
            "types": {"in0": "single", "state": "uint16", "out0": "single"},
        },
        {
            "purpose": "Верхняя граница диапазона.",
            "method": "Граничные значения",
            "actions": f"Вызвать {func}() при in0={hi}.",
            "expect": "Корректная обработка верхней границы.",
            "inputs": {"in0": hi, "state": 0},
            "outputs": {"out0": hi, "state": 1},
            "types": {"in0": "single", "state": "uint16", "out0": "single"},
        },
        {
            "purpose": "Робастность: значение вне диапазона / нештатный вход.",
            "method": "Робастность",
            "actions": f"Вызвать {func}() при in0={hi * 10 if hi else 1e6}.",
            "expect": "Нет переполнения без фиксации; выход ограничен или ошибка зафиксирована.",
            "inputs": {"in0": (hi * 10 if hi else 1e6), "state": 0},
            "outputs": {"out0": hi or 0, "state": 0},
            "types": {"in0": "single", "state": "uint16", "out0": "single"},
        },
        {
            "purpose": "Покрытие ветви альтернативного решения (else / сброс).",
            "method": "Покрытие ветвей",
            "actions": f"Вызвать {func}() при нулевом/инверсном условии относительно типового.",
            "expect": "Проход по альтернативной ветви, счётчики/флаги согласно требованию.",
            "inputs": {"in0": 0, "state": 160},
            "outputs": {"out0": 0, "state": 160},
            "types": {"in0": "single", "state": "uint16", "out0": "single"},
        },
    ]
    return {
        "func": func,
        "danger": danger,
        "hw": hw,
        "cases": cases,
        "mode": "heuristic",
    }


def _llm_suite(req: Requirement, ctx: dict[str, Any]) -> dict[str, Any]:
    llm = chat_llm("expensive")
    if llm is None:
        return _heuristic_suite(req)
    from langchain_core.messages import HumanMessage, SystemMessage

    system = PROMPT.read_text(encoding="utf-8") if PROMPT.exists() else (
        "Инженер модульного тестирования МК. Верни только JSON."
    )
    payload = {
        "requirement": ctx or {"code": req.code, "text": req.text, "attributes": {}},
        "need": [
            "func — имя функции UUT",
            "danger — Низкая или ВЫСОКАЯ: ...",
            "cases[]: purpose, method (Граничные значения|Классы эквивалентности|Робастность|Покрытие ветвей), actions, expect",
            "inputs и outputs — словари имя->число, types имя->single|uint16|int16|int32",
            "не меньше 5 кейсов, покрыть ветви и границы",
        ],
    }
    msg = llm.invoke(
        [
            SystemMessage(content=system + "\nОтвет — один JSON-объект без markdown."),
            HumanMessage(content=json.dumps(payload, ensure_ascii=False)[:8000]),
        ]
    )
    raw = msg.content or "{}"
    m = re.search(r"\{.*\}", raw, re.S)
    data = json.loads(m.group(0) if m else "{}")
    if not data.get("cases"):
        h = _heuristic_suite(req)
        h["mode"] = "llm-fallback"
        return h
    data.setdefault("func", _guess_func(req))
    data.setdefault("danger", "Низкая")
    data["mode"] = "llm"
    return data


def simulate(suite: dict[str, Any], source_code: str | None) -> list[dict[str, Any]]:
    """Прогон. Без кода МК не исполняем — только статический разбор, если код передали."""
    func = suite.get("func") or "uut"
    rows = []
    if not source_code or not source_code.strip():
        for i, c in enumerate(suite["cases"], 1):
            rows.append(
                {
                    "n": i,
                    "purpose": c.get("purpose"),
                    "status": "blocked",
                    "detail": "Код модуля не приложен — моделирование не выполнялось.",
                }
            )
        return rows
    has_fn = bool(re.search(rf"\b{re.escape(func)}\s*\(", source_code))
    regs = bool(re.search(r"\b(RCC|GPIO|ADC|TIM|DMA|NVIC)\w*", source_code))
    for i, c in enumerate(suite["cases"], 1):
        if not has_fn:
            st, det = "blocked", f"В коде нет функции {func}() — кейс не привязан."
        elif "Робастность" in (c.get("method") or "") and regs:
            st, det = "review", "Статика: есть работа с регистрами. Прогон на железе — только после ревью опасности."
        else:
            st, det = "static-ok", f"Функция {func}() найдена. Исполнение на МК не запускалось (нет стенда)."
        rows.append({"n": i, "purpose": c.get("purpose"), "status": st, "detail": det})
    return rows


def write_workbook(req: Requirement, suite: dict[str, Any], sim: list[dict[str, Any]], dest: Path) -> Path:
    from openpyxl import Workbook

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    info = wb.active
    info.title = "Info"
    fname = dest.name
    info.append(["Проект", 'ПО "Управление MK-114"'])
    info.append(["Дата выполнения", date.today().strftime("%d.%m.%Y")])
    info.append(["Исполнитель", "SpecGraph / инженер верификации"])
    info.append(["Среда верификации", "АСДБ.00.36.1900"])
    info.append(["Объект тестирования", suite.get("func") or ""])
    info.append(["Тестовая процедура", "unit / ФПО-АЗПО"])
    info.append(["Тестовый пример", fname])
    info.append(["Требование ОППО", req.code])
    info.append(["Опасность теста", suite.get("danger") or "Низкая"])
    info.append(["Режим генерации", suite.get("mode") or ""])

    cases = suite["cases"]
    # union of input/output keys
    in_keys: list[str] = []
    out_keys: list[str] = []
    types: dict[str, str] = {}
    for c in cases:
        for k, v in (c.get("types") or {}).items():
            types[k] = v
        for k in (c.get("inputs") or {}):
            if k not in in_keys:
                in_keys.append(k)
        for k in (c.get("outputs") or {}):
            if k not in out_keys:
                out_keys.append(k)
    if not in_keys:
        in_keys = ["in0"]
    if not out_keys:
        out_keys = ["out0"]

    ws_in = wb.create_sheet("InputData")
    ws_in.append(in_keys)
    ws_in.append([types.get(k, "single") for k in in_keys])
    for c in cases:
        ins = c.get("inputs") or {}
        ws_in.append([ins.get(k, "") for k in in_keys])

    ws_out = wb.create_sheet("OutputData")
    ws_out.append(out_keys)
    ws_out.append([types.get(k, "single") for k in out_keys])
    for c in cases:
        outs = c.get("outputs") or {}
        ws_out.append([outs.get(k, "") for k in out_keys])

    ws_c = wb.create_sheet("Comments")
    ws_c.append(["Назначение теста", "Метод", "Проводимые действия", "Ожидаемый результат"])
    ws_c.append(["-", "-", "-", "-"])
    for c in cases:
        ws_c.append([c.get("purpose") or "", c.get("method") or "", c.get("actions") or "", c.get("expect") or ""])

    ws_s = wb.create_sheet("Simulation")
    ws_s.append(["№", "Назначение", "Статус прогона", "Комментарий"])
    for r in sim:
        ws_s.append([r["n"], r.get("purpose") or "", r.get("status"), r.get("detail")])

    wb.save(dest)
    return dest


def run_unit_tests(
    db: Session,
    *,
    document_id: int | None = None,
    requirement_id: int | None = None,
    query: str | None = None,
    source_code: str | None = None,
    **_: Any,
) -> dict[str, Any]:
    q = db.query(Requirement).filter(Requirement.is_current.is_(True))
    if requirement_id:
        q = q.filter(Requirement.id == requirement_id)
    elif document_id:
        q = q.filter(Requirement.document_id == document_id)
    reqs = [r for r in q.all() if not (r.extra or {}).get("stub") and not (r.extra or {}).get("appendix")]
    if query:
        key = query.strip()
        reqs = [r for r in reqs if key.lower() in (r.code or "").lower() or key.lower() in (r.text or "").lower()]
    if not reqs:
        return {"count": 0, "files": [], "note": "нет текущих требований под фильтр"}

    EXPORTS.mkdir(parents=True, exist_ok=True)
    files = []
    for req in reqs:
        ctx = expand_requirement(db, req.id)
        suite = _llm_suite(req, ctx)
        sim = simulate(suite, source_code)
        fname = f"M_TEST_{_safe_name(req.code)}_V01 {suite.get('func') or 'uut'}().xlsx"
        path = EXPORTS / fname
        write_workbook(req, suite, sim, path)
        files.append(
            {
                "requirement": req.code,
                "file": str(path),
                "download": f"/exports/{path.name}",
                "danger": suite.get("danger"),
                "func": suite.get("func"),
                "cases": len(suite["cases"]),
                "simulation": sim,
                "mode": suite.get("mode"),
            }
        )
    return {"count": len(files), "files": files, "code_attached": bool(source_code and source_code.strip())}
