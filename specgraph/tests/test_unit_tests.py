from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from specgraph.db import Base
from specgraph.ingest.pipeline import ingest_parsed_json
from specgraph.pipelines.unit_tests import run_unit_tests


def test_unit_xlsx_and_blocked_simulation(tmp_path, monkeypatch):
    from specgraph import config
    from specgraph.pipelines import unit_tests as ut

    monkeypatch.setattr(config.settings, "upload_dir", tmp_path)
    monkeypatch.setattr(ut, "EXPORTS", tmp_path)

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    ingest_parsed_json(
        db,
        {
            "products": [{"code": "P", "name": "МК"}],
            "requirements": [
                {
                    "code": "MK-114.OPPO.HWRQ.054/C",
                    "text": "ПО MCU2 должно инициализировать ADC функцией adcSetup(). Диапазон 0…4095.",
                    "product": "P",
                }
            ],
        },
        index=False,
    )
    out = run_unit_tests(db)
    assert out["count"] == 1
    assert out["code_attached"] is False
    f = Path(out["files"][0]["file"])
    assert f.is_file()
    wb = load_workbook(f)
    assert set(wb.sheetnames) >= {"Info", "InputData", "OutputData", "Comments", "Simulation"}
    assert "HWRQ.054" in str(wb["Info"]["B8"].value) or "HWRQ" in str(wb["Info"]["B8"].value)
    assert wb["Comments"]["B3"].value  # метод
    assert wb["Simulation"]["C2"].value == "blocked"

    out2 = run_unit_tests(db, source_code="void adcSetup(void) { ADC1->CR = 1; }")
    assert out2["code_attached"] is True
    wb2 = load_workbook(out2["files"][0]["file"])
    statuses = [r[2] for r in wb2["Simulation"].iter_rows(min_row=2, values_only=True)]
    assert any(s in {"static-ok", "review"} for s in statuses)
